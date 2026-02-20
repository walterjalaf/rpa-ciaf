"""
Módulo de reporte de estado, errores y telemetría de negocio.

Por qué existe: Centraliza toda la comunicación de resultados al servidor.
El runner reporta éxitos y fallos a través de este módulo, que a su vez
delega en ApiClient. Absorbe errores de telemetría silenciosamente para
que nunca bloqueen el flujo principal.

Uso:
    reporter = Reporter(api_client)
    reporter.report_success(task_id, filepath, date_from, date_to, elapsed)
    reporter.report_failure(task_id, "Error en selector", screenshot_path)
"""

from __future__ import annotations

import logging
from datetime import date
from pathlib import Path

from sync.api_client import ApiClient

logger = logging.getLogger(__name__)


class Reporter:
    """
    Reporta resultados de ejecución al servidor de conciliaciones.

    Por qué existe: Separa la responsabilidad de reportar del runner.
    El runner ejecuta tareas, el reporter envía resultados. Esto permite
    que el reporte falle sin afectar la ejecución de las tareas siguientes.

    Política de errores:
    - report_success: absorbe excepciones con logger.warning
    - report_failure: absorbe excepciones con logger.warning
    - report_session_check: absorbe excepciones con logger.warning

    Nunca propaga: un error en el reporte no debe causar otro error.

    Uso:
        reporter = Reporter(api_client)
        reporter.report_success(task_id, filepath, date_from, date_to, 12.5)
    """

    def __init__(self, api_client: ApiClient) -> None:
        self._api_client = api_client

    def report_success(self, task_id: str, filepath: Path,
                       date_from: date, date_to: date,
                       duration_seconds: float) -> None:
        """
        Reporta la ejecución exitosa de una tarea al servidor.

        Calcula métricas del archivo (tamaño, filas estimadas) y las envía
        como telemetría. Si algo falla, loguea warning pero no propaga.

        Args:
            task_id: Identificador de la tarea completada.
            filepath: Ruta al archivo descargado y subido.
            date_from: Fecha inicio del período descargado.
            date_to: Fecha fin del período descargado.
            duration_seconds: Tiempo de ejecución en segundos.
        """
        try:
            file_size_kb = filepath.stat().st_size / 1024 if filepath.exists() else 0
            row_count = self._estimate_row_count(filepath)

            self._api_client.report_telemetry(
                task_id=task_id,
                date_from=date_from,
                date_to=date_to,
                row_count=row_count,
                file_size_kb=file_size_kb,
                duration_seconds=duration_seconds,
            )
            logger.info(
                "Telemetría enviada: task=%s filas=%d tamaño=%.1fKB",
                task_id, row_count, file_size_kb
            )
        except Exception as e:
            logger.warning("Telemetría no enviada para %s: %s", task_id, e)

    def report_failure(self, task_id: str, error: str,
                       screenshot_path: Path | None = None) -> None:
        """
        Reporta el fallo de una tarea al servidor.

        No propaga excepciones: el reporte de error no debe causar otro error.

        Args:
            task_id: Identificador de la tarea fallida.
            error: Mensaje de error legible.
            screenshot_path: Captura de pantalla del estado del navegador al fallar.
        """
        try:
            self._api_client.report_failure(
                task_id=task_id,
                error=error,
                screenshot_path=screenshot_path,
            )
            logger.info(
                "Fallo reportado: task=%s error='%s'",
                task_id, error[:80]
            )
        except Exception as e:
            logger.warning(
                "No se pudo reportar fallo de %s: %s", task_id, e
            )

    def report_session_check(self, results: list) -> None:
        """
        Envía los resultados del health check al servidor.

        No propaga excepciones: el tracking histórico de sesiones no debe
        bloquear el flujo de la app.

        Args:
            results: Lista de SessionStatus del health checker.
        """
        try:
            self._api_client.report_session_check(results)
            logger.info(
                "Estado de sesiones reportado: %d plataformas",
                len(results)
            )
        except Exception as e:
            logger.warning(
                "No se pudo reportar estado de sesiones: %s", e
            )

    def _estimate_row_count(self, filepath: Path) -> int:
        """
        Estima la cantidad de filas de un archivo Excel o CSV.

        Usa una heurística rápida sin cargar todo el archivo en memoria.
        Si no puede leer el archivo, retorna -1 (valor centinela para
        indicar que no se pudo estimar).
        """
        if not filepath.exists():
            return -1

        try:
            suffix = filepath.suffix.lower()
            if suffix == ".csv":
                # Contar líneas del CSV (rápido, sin parsear)
                with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                    count = sum(1 for _ in f) - 1  # -1 por header
                return max(count, 0)

            if suffix == ".xlsx":
                # Usar openpyxl en modo read_only para no cargar todo
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=True)
                ws = wb.active
                count = ws.max_row - 1 if ws.max_row else 0  # -1 por header
                wb.close()
                return max(count, 0)

        except Exception as e:
            logger.warning(
                "No se pudo estimar filas de %s: %s", filepath.name, e
            )

        return -1


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)

    client = ApiClient()
    client.load_token()
    reporter = Reporter(client)

    # Prueba mock de report_success
    reporter.report_success(
        task_id="mercadopago_movimientos",
        filepath=Path("test.xlsx"),
        date_from=date(2026, 2, 17),
        date_to=date(2026, 2, 17),
        duration_seconds=12.5,
    )

    # Prueba mock de report_failure
    reporter.report_failure(
        task_id="galicia_extracto",
        error="Selector de fecha no encontrado en la página",
    )

    print("Reporter mock: todas las operaciones completadas")
